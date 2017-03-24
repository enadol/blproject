import json
import openpyxl

print "Creating Excel file..."
fh=open('../json/tablavoll16.json')
stringdata=fh.read()
jsondata=json.loads(stringdata)
lst=list()

tabla=jsondata['children']

for i in range(0, 18):
    lugar=tabla[i]['id']
    jornada=tabla[i]['jornada']
    equipo=tabla[i]['name']
    pj=tabla[i]['jugados']
    pg=tabla[i]['ganados']
    pe=tabla[i]['empatados']
    pp=tabla[i]['perdidos']
    gf=tabla[i]['gf']
    gc=tabla[i]['gc']
    dif=tabla[i]['dif']
    puntos=tabla[i]['size']
    lst.append([lugar, equipo, pj, pg, pe, pp, gf, gc, dif, puntos])
#optional
print lst

from openpyxl import Workbook, load_workbook
wb=Workbook()
ws=wb.active

#cell_range=ws['A1':'J1']
#print cell_range

#for cell in cell_range:
ws.append(["ID", "Equipos", "PJ", "PG", "PE", "PP", "GF", "GC", "DIF", "Puntos"])
for item in lst:
    ws.append(item)
    #print dato
            
wb.save('2016.xlsx')
